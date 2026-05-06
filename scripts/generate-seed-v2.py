#!/usr/bin/env python3
"""Generate seed SQL for new/updated taxonomy data from CRA-76 Draft XLSX."""

import re
import openpyxl

XLSX_PATH = '/Users/craefto/Downloads/JapaTak-Taxonomy-CRA76-Draft.xlsx'

def slugify(text):
    if not text:
        return ''
    s = text.lower().strip()
    s = re.sub(r'[^a-z0-9\s-]', '', s)
    s = re.sub(r'[\s]+', '-', s)
    s = re.sub(r'-+', '-', s)
    return s.strip('-')

def esc(val):
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"

def parse_priority(raw):
    if not raw:
        return 'out_of_scope'
    r = str(raw).strip()
    if r.startswith('P1'): return 'P1'
    if r.startswith('P2'): return 'P2'
    if r.startswith('P3'): return 'P3'
    if r.startswith('P4'): return 'P4'
    return 'out_of_scope'

def priority_to_status(p):
    if p == 'P1': return 'active'
    if p == 'out_of_scope': return 'inactive'
    return 'later'

def parse_content_priority(raw):
    if not raw: return None
    r = str(raw).strip().lower()
    if r == 'high': return 'high'
    if r == 'low': return 'low'
    return 'medium'

def parse_region_type(raw):
    if not raw: return None
    mapping = {
        'ski resort': 'ski_resort', 'ski resort suburb': 'ski_resort',
        'rural town': 'rural_town', 'coastal': 'coastal',
        'onsen town': 'onsen_town', 'hot spring + ski': 'onsen_town',
        'urban access': 'urban_access', 'urban/investment': 'urban_access',
        'urban/commercial': 'urban_access', 'cultural/urban': 'urban_access',
        'cultural/heritage': 'rural_town', 'mountain village': 'mountain_village',
        'lakeside': 'lakeside', 'beach/resort': 'coastal',
    }
    return mapping.get(str(raw).strip().lower())

def parse_bool(raw):
    if not raw: return None
    r = str(raw).strip().upper()
    if r in ('Y', 'YES', 'TRUE', '1'): return True
    if r in ('N', 'NO', 'FALSE', '0'): return False
    return None

def parse_int(raw):
    if raw is None: return None
    try:
        s = str(raw).strip()
        m = re.match(r'(\d+)', s)
        return int(m.group(1)) if m else None
    except: return None

def parse_currency(val):
    if val is None: return None
    s = str(val).strip().replace('¥', '').replace('A$', '').replace(',', '').replace('+', '')
    try: return int(float(s))
    except: return None

def map_renovation_category(raw):
    if not raw: return 'interior'
    mapping = {
        'wet areas': 'wet_areas', 'interior': 'interior', 'exterior': 'exterior',
        'maintenance & safety': 'maintenance_safety', 'maintenance safety': 'maintenance_safety',
        'comfort & performance': 'comfort_performance', 'comfort performance': 'comfort_performance',
        'unique to japan': 'unique_to_japan',
    }
    return mapping.get(str(raw).strip().lower(), 'interior')

def map_quiz_weight(raw):
    if not raw: return 'medium'
    r = str(raw).strip().lower()
    if 'high' in r: return 'high'
    if 'low' in r: return 'low'
    return 'medium'

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
lines = []

# === UPDATE CITIES with new columns ===
ws = wb['Geographic Hierarchy']
lines.append('-- Update cities with content_priority, shuttle_bus, slope access')
seen_cities = set()
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0] or not row[2]: continue
    city_slug = slugify(str(row[2]).strip())
    if city_slug in seen_cities: continue
    seen_cities.add(city_slug)

    cp = parse_content_priority(row[8])
    shuttle = parse_bool(row[15])
    bus_slope = parse_int(row[16])
    car_slope = parse_int(row[17])

    sets = []
    if cp: sets.append(f"content_priority = '{cp}'")
    if shuttle is not None: sets.append(f"shuttle_bus = {str(shuttle).lower()}")
    if bus_slope is not None: sets.append(f"bus_to_slope_min = {bus_slope}")
    if car_slope is not None: sets.append(f"car_to_slope_min = {car_slope}")

    if sets:
        lines.append(f"UPDATE cities SET {', '.join(sets)} WHERE slug = {esc(city_slug)};")

lines.append('')

# === SEED LOCAL AREAS ===
lines.append('-- Local Areas')
seen_areas = set()
area_order = 0
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[4] or not row[2]: continue
    city_slug = slugify(str(row[2]).strip())
    area_en = str(row[4]).strip()
    area_ja = str(row[5]).strip() if row[5] else ''
    area_slug = slugify(area_en)
    key = (city_slug, area_slug)
    if key in seen_areas: continue
    seen_areas.add(key)
    area_order += 1

    rt = parse_region_type(row[6])
    rt_sql = f"'{rt}'" if rt else 'NULL'

    lines.append(
        f"INSERT INTO local_areas (slug, name_en, name_ja, city_id, region_type, sort_order) "
        f"VALUES ({esc(area_slug)}, {esc(area_en)}, {esc(area_ja)}, "
        f"(SELECT id FROM cities WHERE slug = {esc(city_slug)}), "
        f"{rt_sql}, {area_order});"
    )
lines.append('')

# === SEED PROPERTY CONDITIONS ===
ws = wb['Property Condition']
lines.append('-- Property Conditions')
pc_order = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[0] or not str(row[0]).startswith('PC'): continue
    pc_order += 1
    name_en = str(row[1]).strip()
    name_ja = str(row[2]).strip() if row[2] else ''
    desc = str(row[3]).strip() if row[3] else ''
    target = str(row[4]).strip() if row[4] else ''
    weight = map_quiz_weight(row[5])
    notes = str(row[6]).strip() if row[6] else ''

    lines.append(
        f"INSERT INTO property_conditions (slug, name_en, name_ja, description, target_persona, "
        f"quiz_weight, notes, sort_order) "
        f"VALUES ({esc(slugify(name_en))}, {esc(name_en)}, {esc(name_ja)}, "
        f"{esc(desc)}, {esc(target)}, '{weight}', {esc(notes)}, {pc_order});"
    )
lines.append('')

# === RE-SEED RENOVATION FEATURES (clean 34 with buyer_relevance) ===
ws = wb['Renovation Features']
lines.append('-- Renovation Features (clean 34 with buyer_relevance)')
rf_order = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[0] or not str(row[0]).startswith('RF'): continue
    rf_order += 1
    name_en = str(row[1]).strip()
    name_ja = str(row[2]).strip() if row[2] else ''
    category = map_renovation_category(row[3])
    desc = str(row[4]).strip() if row[4] else ''
    relevance = str(row[5]).strip() if row[5] else ''
    japan_specific = str(row[6]).strip().lower().startswith('yes') if row[6] else False

    lines.append(
        f"INSERT INTO renovation_features (slug, name_en, name_ja, category, description, "
        f"buyer_relevance, japan_specific, status, sort_order) "
        f"VALUES ({esc(slugify(name_en))}, {esc(name_en)}, {esc(name_ja)}, "
        f"'{category}', {esc(desc)}, {esc(relevance)}, "
        f"{'true' if japan_specific else 'false'}, 'active', {rf_order});"
    )

sql = '\n'.join(lines)
print(sql)
print(f'\n-- Summary: {len(seen_areas)} local areas, {pc_order} property conditions, {rf_order} renovation features')
